"""Módulo de estilos para reportes PDF y Excel.

Define paleta de colores, estilos tipográficos y funciones auxiliares
para generar reportes con formato consistente tanto en PDF (via ReportLab)
como en Excel (via OpenPyXL).
"""

from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter

# Paleta de colores corporativos
COLOR_PRIMARIO = '#0f172a'
COLOR_ACENTO = '#3b82f6'
COLOR_BORDER = '#e2e8f0'
COLOR_ALT_FILA = '#f8fafc'
COLOR_TEXTO_MUTED = '#64748b'


def estilos_pdf():
    """Devuelve un diccionario con estilos de párrafo para reportes PDF."""
    return {
        'titulo': ParagraphStyle(
            'TituloReporte', fontSize=18, leading=22,
            textColor=colors.HexColor(COLOR_PRIMARIO), spaceAfter=4,
            fontName='Helvetica-Bold', alignment=TA_LEFT
        ),
        'subtitulo': ParagraphStyle(
            'SubtituloReporte', fontSize=9, leading=12,
            textColor=colors.HexColor(COLOR_TEXTO_MUTED), spaceAfter=16,
            fontName='Helvetica', alignment=TA_LEFT
        ),
        'celda': ParagraphStyle(
            'CeldaReporte', fontSize=9, leading=12,
            textColor=colors.HexColor(COLOR_PRIMARIO), fontName='Helvetica'
        ),
        'celda_bold': ParagraphStyle(
            'CeldaBoldReporte', fontSize=9, leading=12,
            textColor=colors.HexColor(COLOR_PRIMARIO), fontName='Helvetica-Bold'
        ),
        'header': ParagraphStyle(
            'HeaderReporte', fontSize=9, leading=12,
            textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER
        ),
    }


def estilo_tabla_pdf(num_filas):
    """Crea y devuelve un TableStyle para tablas PDF con alternancia de colores en filas."""
    # Estilos base para encabezado, cuerpo y cuadrícula
    s = TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  colors.HexColor(COLOR_PRIMARIO)),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING',  (0, 0), (-1, 0),  10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING',(0, 0), (-1, -1), 10),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 1), (-1, -1), 9),
        ('TEXTCOLOR',   (0, 1), (-1, -1), colors.HexColor(COLOR_PRIMARIO)),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING',  (0, 1), (-1, -1), 8),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
        ('LINEBELOW',   (0, 0), (-1, 0),  1.5, colors.HexColor('#1e293b')),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
    ])
    # Alternar color de fondo en filas pares (filas de datos)
    for i in range(1, num_filas + 1, 2):
        s.add('BACKGROUND', (0, i), (-1, i), colors.HexColor(COLOR_ALT_FILA))
    return s


def build_pdf_tabla(encabezados, filas, estilos, ancho_disponible, columnas_centradas=None):
    """Construye y devuelve un objeto Table de ReportLab listo para insertar en PDF."""
    # Encabezados renderizados como párrafos con estilo header
    tabla_data = [[Paragraph(h, estilos['header']) for h in encabezados]]
    # Filas de datos: valores None se muestran como guion largo
    for fila in filas:
        tabla_data.append([Paragraph(str(v) if v is not None else '\u2014', estilos['celda']) for v in fila])
    ncols = len(encabezados)
    col_widths = [ancho_disponible / ncols] * ncols
    table = Table(tabla_data, colWidths=col_widths)
    table.setStyle(estilo_tabla_pdf(len(filas)))
    # Centrar columnas específicas si se solicita
    if columnas_centradas:
        for col in columnas_centradas:
            table.setStyle(TableStyle([('ALIGN', (col, 1), (col, -1), 'CENTER')]))
    return table


def configurar_encabezado_excel(ws, columnas, titulo=None):
    """Configura el encabezado de una hoja Excel: estilos, título opcional y congelación de paneles.

    Retorna la fila donde deben empezar los datos.
    """
    # Estilos del encabezado
    header_fill = PatternFill(start_color=COLOR_PRIMARIO.lstrip('#'),
                              end_color=COLOR_PRIMARIO.lstrip('#'), fill_type='solid')
    header_font = Font(name='Segoe UI', bold=True, color='ffffff', size=10)
    header_border = Border(bottom=Side(style='medium', color='1e293b'))

    if titulo:
        # Fusionar celdas para el título y aplicar estilo grande
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = Font(name='Segoe UI', bold=True, size=16, color=COLOR_PRIMARIO.lstrip('#'))
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        header_row = 2
        data_start = 3
    else:
        header_row = 1
        data_start = 2

    # Aplicar estilos a cada columna del encabezado
    for idx, nombre in enumerate(columnas, 1):
        cel = ws.cell(row=header_row, column=idx, value=nombre)
        cel.fill = header_fill
        cel.font = header_font
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border = header_border
    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return data_start


def aplicar_estilo_datos_excel(ws, columnas, data_start):
    """Aplica estilos al cuerpo de datos en Excel: fuente, bordes, alternancia de color y auto-ancho."""
    # Estilos para las celdas de datos
    body_font = Font(name='Segoe UI', size=10, color=COLOR_PRIMARIO.lstrip('#'))
    alt_fill = PatternFill(start_color=COLOR_ALT_FILA.lstrip('#'),
                           end_color=COLOR_ALT_FILA.lstrip('#'), fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color=COLOR_BORDER.lstrip('#')))

    if ws.max_row < data_start:
        return

    # Aplicar fuente, bordes y color alterno por fila
    for r_idx, fila in enumerate(ws.iter_rows(min_row=data_start,
                                               max_row=ws.max_row,
                                               max_col=len(columnas))):
        for celda in fila:
            celda.font = body_font
            celda.border = thin_border
            celda.alignment = Alignment(vertical='center')
            if (r_idx % 2) == 1:
                celda.fill = alt_fill

    # Ajustar ancho de columna según el contenido más largo (máx. 45)
    for col_idx in range(1, len(columnas) + 1):
        max_len = len(str(columnas[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=data_start, values_only=True):
            for val in row:
                if val is not None:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 5, 45)
