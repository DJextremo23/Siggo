"""
Blueprint para reportes del fiscalizador - resumen personal, detalle y vacaciones con exportación a PDF y Excel
"""
from flask import Blueprint, render_template, request, send_file, session, redirect, url_for
from io import BytesIO
from conexion import conexion
from utils import acceso_no_autorizado
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from datetime import datetime, date
from estilos_reporte import (
    COLOR_PRIMARIO, COLOR_ACENTO, COLOR_TEXTO_MUTED,
    estilos_pdf, build_pdf_tabla,
    configurar_encabezado_excel, aplicar_estilo_datos_excel
)


def _estilos_pdf():
    return estilos_pdf()


def _build_pdf_tabla(encabezados, filas, estilos, ancho_disponible, columnas_centradas=None):
    return build_pdf_tabla(encabezados, filas, estilos, ancho_disponible, columnas_centradas)


def _configurar_encabezado_excel(ws, columnas, titulo=None):
    return configurar_encabezado_excel(ws, columnas, titulo)


def _aplicar_estilo_datos_excel(ws, columnas, data_start):
    aplicar_estilo_datos_excel(ws, columnas, data_start)


# Creación del blueprint para las rutas de mis reportes
mis_reportes_bp = Blueprint("mis_reportes_bp", __name__)


# Ruta principal: tabla resumen + detalle de guardias + vacaciones del fiscalizador
@mis_reportes_bp.route("/reportes")
def mis_reportes():

    if "usuario" not in session:
        return redirect(url_for("home"))

    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()

    id_usuario = session.get("id_usuario")

    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    buscar = request.args.get("buscar", "").strip()
    buscar_vac = request.args.get("buscar_vac", "").strip()

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        # ================= FILTRO BASE =================
        filtro_base = " WHERE g.id_usuario = %s "
        params_base = [id_usuario]

        if anio:
            filtro_base += " AND YEAR(g.fecha_guardia) = %s "
            params_base.append(int(anio))

        if fecha_desde:
            filtro_base += " AND g.fecha_guardia >= %s "
            params_base.append(fecha_desde)

        if fecha_hasta:
            filtro_base += " AND g.fecha_guardia <= %s "
            params_base.append(fecha_hasta)

        filtro_detalle = filtro_base
        params_detalle = params_base.copy()

        if buscar:
            like = f"%{buscar}%"
            filtro_detalle += """ AND (
                g.fecha_guardia LIKE %s OR
                COALESCE(f.descripcion,'') LIKE %s OR
                a.estado LIKE %s OR
                CASE WHEN a.estado='asistio' THEN '✔ Asistió'
                     WHEN a.estado='falta' THEN '❌ Falta'
                     WHEN a.estado='justificado' THEN '🟡 Justificado'
                     ELSE '—' END LIKE %s OR
                CASE WHEN a.estado='falta' THEN '❌ No cumple'
                     WHEN a.estado='justificado' THEN '🟡 Justificado'
                     WHEN a.estado='asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                     WHEN a.estado='asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                     ELSE '—' END LIKE %s
            ) """
            params_detalle.extend([like, like, like, like, like])

        # ================= RESUMEN =================
        cursor.execute(f"""
            SELECT
                COUNT(g.id_guardia) AS total_guardias,

                SUM(CASE WHEN f.id_feriado IS NOT NULL THEN 1 ELSE 0 END)
                AS guardias_feriado,

                COUNT(c.id_compensacion) AS compensaciones,

                (COUNT(g.id_guardia) - COUNT(c.id_compensacion)) AS pendientes

            FROM guardias g

            LEFT JOIN feriados f
                ON g.id_feriado = f.id_feriado

            LEFT JOIN compensaciones c
                ON g.id_guardia = c.id_guardia

            {filtro_base}
        """, params_base)

        reporte = cursor.fetchall()

        # ================= DETALLE =================
        cursor.execute(f"""
            SELECT
                g.fecha_guardia,
                COALESCE(f.descripcion,'—') AS feriado,

                CASE
                    WHEN a.estado='asistio' THEN '✔ Asistió'
                    WHEN a.estado='falta' THEN '❌ Falta'
                    WHEN a.estado='justificado' THEN '🟡 Justificado'
                    ELSE '—'
                END AS asistencia,

                c.fecha_compensacion,
                c.observacion,

                CASE
                    WHEN a.estado='falta' THEN '❌ No cumple'
                    WHEN a.estado='justificado' THEN '🟡 Justificado'
                    WHEN a.estado='asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                    WHEN a.estado='asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                    ELSE '—'
                END AS estado_compensacion

            FROM guardias g

            LEFT JOIN asistencia a
                ON g.id_guardia = a.id_guardia

            LEFT JOIN compensaciones c
                ON g.id_guardia = c.id_guardia

            LEFT JOIN feriados f
                ON g.id_feriado = f.id_feriado

            {filtro_detalle}

            ORDER BY g.fecha_guardia DESC
        """, params_detalle)

        detalle = cursor.fetchall()

        # ================= VACACIONES =================
        filtro_vac = "WHERE v.id_usuario = %s"
        params_vac = [id_usuario]

        if anio:
            filtro_vac += " AND (YEAR(v.fecha_inicio) = %s OR YEAR(v.fecha_fin) = %s)"
            params_vac.extend([int(anio), int(anio)])

        if fecha_desde:
            filtro_vac += " AND v.fecha_fin >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_inicio <= %s"
            params_vac.append(fecha_hasta)

        if buscar_vac:
            filtro_vac += " AND (v.fecha_inicio LIKE %s OR v.fecha_fin LIKE %s OR v.estado LIKE %s) "
            like_vac = f"%{buscar_vac}%"
            params_vac.extend([like_vac, like_vac, like_vac])

        cursor.execute(f"""
            SELECT fecha_inicio, fecha_fin, estado
            FROM vacaciones v
            {filtro_vac}
            ORDER BY v.fecha_inicio DESC
        """, params_vac)

        vacaciones = cursor.fetchall()

        dias_tomados = 0
        for v in vacaciones:
            if v["fecha_inicio"] and v["fecha_fin"]:
                dias_tomados += (v["fecha_fin"] - v["fecha_inicio"]).days + 1

        cursor.execute(
            "SELECT fecha_ingreso FROM usuarios WHERE id_usuario = %s",
            (id_usuario,)
        )
        user = cursor.fetchone()
        dias_faltantes = 0
        if user and user.get("fecha_ingreso"):
            fecha_ingreso = user["fecha_ingreso"]
            today = date.today()
            anniv = fecha_ingreso.replace(year=fecha_ingreso.year + 1)
            years = 0
            while anniv <= today:
                years += 1
                anniv = anniv.replace(year=anniv.year + 1)
            total_dias = years * 30

            cursor.execute("""
                SELECT COALESCE(SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1), 0) AS total
                FROM vacaciones v
                WHERE v.id_usuario = %s
            """, (id_usuario,))
            result = cursor.fetchone()
            total_tomados = result["total"] if result else 0
            dias_faltantes = max(0, total_dias - total_tomados)

    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    return render_template(
        "mis_reportes.html",
        reporte=reporte,
        detalle=detalle,
        vacaciones=vacaciones,
        dias_tomados=dias_tomados,
        dias_faltantes=dias_faltantes
    )


# ==========================
# EXPORTAR PDF (DETALLE)
# ==========================
# Exportación a PDF del detalle de guardias
@mis_reportes_bp.route("/exportar/pdf")
def exportar_pdf():
    if "usuario" not in session:
        return redirect(url_for("home"))

    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()

    id_usuario = session["id_usuario"]
    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    buscar = request.args.get("buscar", "").strip()

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_detalle = "WHERE g.id_usuario = %s"
        params_detalle = [id_usuario]

        if anio:
            anio_int = int(anio)
            filtro_detalle += " AND YEAR(g.fecha_guardia) = %s"
            params_detalle.append(anio_int)

        if fecha_desde:
            filtro_detalle += " AND g.fecha_guardia >= %s"
            params_detalle.append(fecha_desde)

        if fecha_hasta:
            filtro_detalle += " AND g.fecha_guardia <= %s"
            params_detalle.append(fecha_hasta)

        if buscar:
            like = f"%{buscar}%"
            filtro_detalle += """ AND (
                g.fecha_guardia LIKE %s OR
                COALESCE(f.descripcion,'') LIKE %s OR
                a.estado LIKE %s OR
                CASE WHEN a.estado='asistio' THEN '✔ Asistió'
                     WHEN a.estado='falta' THEN '❌ Falta'
                     WHEN a.estado='justificado' THEN '🟡 Justificado'
                     ELSE '—' END LIKE %s OR
                CASE WHEN a.estado='falta' THEN '❌ No cumple'
                     WHEN a.estado='justificado' THEN '🟡 Justificado'
                     WHEN a.estado='asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                     WHEN a.estado='asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                     ELSE '—' END LIKE %s
            ) """
            params_detalle.extend([like, like, like, like, like])

        cursor.execute(f"""
            SELECT
                g.fecha_guardia,
                COALESCE(f.descripcion,'—') AS feriado,

                CASE
                    WHEN a.estado='asistio' THEN '✔ Asistió'
                    WHEN a.estado='falta' THEN '❌ Falta'
                    WHEN a.estado='justificado' THEN '🟡 Justificado'
                    ELSE '—'
                END AS asistencia,

                c.fecha_compensacion,
                c.observacion,

                CASE
                    WHEN a.estado='falta' THEN '❌ No cumple'
                    WHEN a.estado='justificado' THEN '🟡 Justificado'
                    WHEN a.estado='asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                    WHEN a.estado='asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                    ELSE '—'
                END AS estado_compensacion

            FROM guardias g
            LEFT JOIN asistencia a ON g.id_guardia=a.id_guardia
            LEFT JOIN feriados f ON g.id_feriado=f.id_feriado
            LEFT JOIN compensaciones c ON g.id_guardia=c.id_guardia
            {filtro_detalle}
            ORDER BY g.fecha_guardia DESC
        """, params_detalle)

        detalle = cursor.fetchall()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=25, rightMargin=25, topMargin=30, bottomMargin=30)
    elementos = []

    elementos.append(Paragraph("Detalle de Guardias", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Fecha", "Feriado", "Asistencia", "Compensación", "Estado", "Observaciones"]
    filas = [[
        str(d["fecha_guardia"]) if d["fecha_guardia"] else "",
        d["feriado"] or "\u2014",
        d["asistencia"] or "\u2014",
        str(d["fecha_compensacion"]) if d["fecha_compensacion"] else "\u2014",
        d["estado_compensacion"] or "\u2014",
        d["observacion"] or "\u2014"
    ] for d in detalle]
    ancho = landscape(A4)[0] - 50

    table = _build_pdf_tabla(columnas, filas, estilos, ancho)

    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, download_name="mis_reportes.pdf", as_attachment=True)


# ==========================
# EXPORTAR EXCEL (DETALLE)
# ==========================
# Exportación a Excel del detalle de guardias
@mis_reportes_bp.route("/exportar/excel")
def exportar_excel():

    if "usuario" not in session:
        return redirect(url_for("home"))

    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()

    id_usuario = session["id_usuario"]
    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    buscar = request.args.get("buscar", "").strip()

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_detalle = "WHERE g.id_usuario = %s"
        params_detalle = [id_usuario]

        if anio:
            anio_int = int(anio)
            filtro_detalle += " AND YEAR(g.fecha_guardia) = %s"
            params_detalle.append(anio_int)

        if fecha_desde:
            filtro_detalle += " AND g.fecha_guardia >= %s"
            params_detalle.append(fecha_desde)

        if fecha_hasta:
            filtro_detalle += " AND g.fecha_guardia <= %s"
            params_detalle.append(fecha_hasta)

        if buscar:
            like = f"%{buscar}%"
            filtro_detalle += """ AND (
                g.fecha_guardia LIKE %s OR
                COALESCE(f.descripcion,'') LIKE %s OR
                a.estado LIKE %s OR
                CASE WHEN a.estado='asistio' THEN '✔ Asistió'
                     WHEN a.estado='falta' THEN '❌ Falta'
                     WHEN a.estado='justificado' THEN '🟡 Justificado'
                     ELSE '—' END LIKE %s OR
                CASE WHEN a.estado='falta' THEN '❌ No cumple'
                     WHEN a.estado='justificado' THEN '🟡 Justificado'
                     WHEN a.estado='asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                     WHEN a.estado='asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                     ELSE '—' END LIKE %s
            ) """
            params_detalle.extend([like, like, like, like, like])

        cursor.execute(f"""
            SELECT
                g.fecha_guardia,
                COALESCE(f.descripcion,'—') AS feriado,

                CASE
                    WHEN a.estado='asistio' THEN '✔ Asistió'
                    WHEN a.estado='falta' THEN '❌ Falta'
                    WHEN a.estado='justificado' THEN '🟡 Justificado'
                    ELSE '—'
                END AS asistencia,

                c.fecha_compensacion,
                c.observacion,

                CASE
                    WHEN a.estado='falta' THEN '❌ No cumple'
                    WHEN a.estado='justificado' THEN '🟡 Justificado'
                    WHEN a.estado='asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                    WHEN a.estado='asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                    ELSE '—'
                END AS estado_compensacion

            FROM guardias g
            LEFT JOIN asistencia a ON g.id_guardia = a.id_guardia
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            {filtro_detalle}
            ORDER BY g.fecha_guardia DESC
        """, params_detalle)

        detalle = cursor.fetchall()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"
    columnas = ["Fecha", "Feriado", "Asistencia", "Compensación", "Estado", "Observaciones"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Detalle de Guardias")
    for d in detalle:
        ws.append([
            str(d["fecha_guardia"]) if d["fecha_guardia"] else "",
            d["feriado"] or "",
            d["asistencia"] or "",
            str(d["fecha_compensacion"]) if d["fecha_compensacion"] else "",
            d["estado_compensacion"] or "",
            d["observacion"] or ""
        ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="mis_reportes.xlsx",
                     as_attachment=True)


# ==========================
# EXPORTAR PDF (RESUMEN)
# ==========================
# Exportación a PDF del resumen de guardias
@mis_reportes_bp.route("/exportar/resumen/pdf")
def exportar_resumen_pdf():
    if "usuario" not in session:
        return redirect(url_for("home"))
    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()
    id_usuario = session["id_usuario"]
    nombre = session.get("nombre", "—")
    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_base = " WHERE g.id_usuario = %s "
        params_base = [id_usuario]
        if anio:
            filtro_base += " AND YEAR(g.fecha_guardia) = %s "
            params_base.append(int(anio))
        if fecha_desde:
            filtro_base += " AND g.fecha_guardia >= %s "
            params_base.append(fecha_desde)
        if fecha_hasta:
            filtro_base += " AND g.fecha_guardia <= %s "
            params_base.append(fecha_hasta)

        cursor.execute(f"""
            SELECT
                COUNT(g.id_guardia) AS total_guardias,
                SUM(CASE WHEN f.id_feriado IS NOT NULL THEN 1 ELSE 0 END) AS guardias_feriado,
                COUNT(c.id_compensacion) AS compensaciones,
                (COUNT(g.id_guardia) - COUNT(c.id_compensacion)) AS pendientes
            FROM guardias g
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            {filtro_base}
        """, params_base)
        resumen = cursor.fetchall()
        r = resumen[0] if resumen else {}
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elementos = []

    elementos.append(Paragraph("Resumen de Guardias", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    partes.append(f"Fiscalizador: {nombre}")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Fiscalizador", "Total Guardia", "Total Feriado", "Total Compensaciones", "Compensaciones Pendientes"]
    filas = [[
        nombre,
        str(r.get("total_guardias", 0)),
        str(r.get("guardias_feriado", 0)),
        str(r.get("compensaciones", 0)),
        str(r.get("pendientes", 0))
    ]]
    ancho = landscape(A4)[0] - 60

    table = _build_pdf_tabla(columnas, filas, estilos, ancho, columnas_centradas=[1, 2, 3, 4])

    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)
    return send_file(buffer, download_name="resumen_guardias.pdf", as_attachment=True)


# ==========================
# EXPORTAR EXCEL (RESUMEN)
# ==========================
# Exportación a Excel del resumen de guardias
@mis_reportes_bp.route("/exportar/resumen/excel")
def exportar_resumen_excel():
    if "usuario" not in session:
        return redirect(url_for("home"))
    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()
    id_usuario = session["id_usuario"]
    nombre = session.get("nombre", "—")
    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_base = " WHERE g.id_usuario = %s "
        params_base = [id_usuario]
        if anio:
            filtro_base += " AND YEAR(g.fecha_guardia) = %s "
            params_base.append(int(anio))
        if fecha_desde:
            filtro_base += " AND g.fecha_guardia >= %s "
            params_base.append(fecha_desde)
        if fecha_hasta:
            filtro_base += " AND g.fecha_guardia <= %s "
            params_base.append(fecha_hasta)

        cursor.execute(f"""
            SELECT
                COUNT(g.id_guardia) AS total_guardias,
                SUM(CASE WHEN f.id_feriado IS NOT NULL THEN 1 ELSE 0 END) AS guardias_feriado,
                COUNT(c.id_compensacion) AS compensaciones,
                (COUNT(g.id_guardia) - COUNT(c.id_compensacion)) AS pendientes
            FROM guardias g
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            {filtro_base}
        """, params_base)
        resumen = cursor.fetchall()
        r = resumen[0] if resumen else {}
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    columnas = ["Fiscalizador", "Total Guardia", "Total Feriado", "Total Compensaciones", "Compensaciones Pendientes"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Resumen de Guardias")
    ws.append([
        nombre,
        r.get("total_guardias", 0),
        r.get("guardias_feriado", 0),
        r.get("compensaciones", 0),
        r.get("pendientes", 0)
    ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=2, max_col=5):
        for cel in row:
            cel.alignment = Alignment(horizontal='center', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="resumen_guardias.xlsx", as_attachment=True)


# ==========================
# EXPORTAR PDF (VACACIONES)
# ==========================
# Exportación a PDF de vacaciones del fiscalizador
@mis_reportes_bp.route("/exportar/vacaciones/pdf")
def exportar_vacaciones_pdf():

    if "usuario" not in session:
        return redirect(url_for("home"))

    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()

    id_usuario = session["id_usuario"]
    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    buscar_vac = request.args.get("buscar_vac", "").strip()

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_vac = "WHERE v.id_usuario = %s"
        params_vac = [id_usuario]

        if anio:
            filtro_vac += " AND (YEAR(v.fecha_inicio) = %s OR YEAR(v.fecha_fin) = %s)"
            params_vac.extend([int(anio), int(anio)])

        if fecha_desde:
            filtro_vac += " AND v.fecha_fin >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_inicio <= %s"
            params_vac.append(fecha_hasta)

        if buscar_vac:
            filtro_vac += " AND (v.fecha_inicio LIKE %s OR v.fecha_fin LIKE %s OR v.estado LIKE %s) "
            like_vac = f"%{buscar_vac}%"
            params_vac.extend([like_vac, like_vac, like_vac])

        cursor.execute(f"""
            SELECT fecha_inicio, fecha_fin, estado
            FROM vacaciones v
            {filtro_vac}
            ORDER BY v.fecha_inicio DESC
        """, params_vac)

        data = cursor.fetchall()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    dias_tomados = 0
    for v in data:
        if v["fecha_inicio"] and v["fecha_fin"]:
            dias_tomados += (v["fecha_fin"] - v["fecha_inicio"]).days + 1
    dias_faltantes = max(0, 30 - dias_tomados)

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elementos = []

    elementos.append(Paragraph("Detalle de Vacaciones", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    partes.append(f"Tomados: {dias_tomados} días")
    partes.append(f"Faltantes: {dias_faltantes} de 30 días")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Inicio", "Fin", "Días", "Estado"]
    filas = [[
        str(v["fecha_inicio"]) if v["fecha_inicio"] else "\u2014",
        str(v["fecha_fin"]) if v["fecha_fin"] else "\u2014",
        str((v["fecha_fin"] - v["fecha_inicio"]).days + 1) if v["fecha_inicio"] and v["fecha_fin"] else "\u2014",
        v["estado"].replace('_', ' ').title() if v["estado"] else "\u2014"
    ] for v in data]
    ancho = landscape(A4)[0] - 60

    table = _build_pdf_tabla(columnas, filas, estilos, ancho)

    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer,
                     download_name="vacaciones.pdf",
                     as_attachment=True)


# ==========================
# EXPORTAR EXCEL (VACACIONES)
# ==========================
# Exportación a Excel de vacaciones del fiscalizador
@mis_reportes_bp.route("/exportar/vacaciones/excel")
def exportar_vacaciones_excel():

    if "usuario" not in session:
        return redirect(url_for("home"))

    if session.get("perfil_activo") != "fiscalizador":
        return acceso_no_autorizado()

    id_usuario = session["id_usuario"]
    anio = request.args.get("anio")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    buscar_vac = request.args.get("buscar_vac", "").strip()

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_vac = "WHERE v.id_usuario = %s"
        params_vac = [id_usuario]

        if anio:
            filtro_vac += " AND (YEAR(v.fecha_inicio) = %s OR YEAR(v.fecha_fin) = %s)"
            params_vac.extend([int(anio), int(anio)])

        if fecha_desde:
            filtro_vac += " AND v.fecha_fin >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_inicio <= %s"
            params_vac.append(fecha_hasta)

        if buscar_vac:
            filtro_vac += " AND (v.fecha_inicio LIKE %s OR v.fecha_fin LIKE %s OR v.estado LIKE %s) "
            like_vac = f"%{buscar_vac}%"
            params_vac.extend([like_vac, like_vac, like_vac])

        cursor.execute(f"""
            SELECT fecha_inicio, fecha_fin, estado
            FROM vacaciones v
            {filtro_vac}
            ORDER BY v.fecha_inicio DESC
        """, params_vac)

        data = cursor.fetchall()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    dias_tomados = 0
    for v in data:
        if v["fecha_inicio"] and v["fecha_fin"]:
            dias_tomados += (v["fecha_fin"] - v["fecha_inicio"]).days + 1
    dias_faltantes = max(0, 30 - dias_tomados)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vacaciones"
    columnas = ["Inicio", "Fin", "Días", "Estado"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Detalle de Vacaciones")

    ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=len(columnas))
    c = ws.cell(row=data_start, column=1)
    c.value = f"Tomados: {dias_tomados} días  |  Faltantes: {dias_faltantes} de 30 días"
    c.font = Font(name='Segoe UI', italic=True, size=9, color=COLOR_TEXTO_MUTED.lstrip('#'))
    c.alignment = Alignment(horizontal='left', vertical='center')
    data_start += 1

    for v in data:
        ws.append([
            str(v["fecha_inicio"]) if v["fecha_inicio"] else "",
            str(v["fecha_fin"]) if v["fecha_fin"] else "",
            (v["fecha_fin"] - v["fecha_inicio"]).days + 1 if v["fecha_inicio"] and v["fecha_fin"] else "",
            v["estado"].replace('_', ' ').title() if v["estado"] else ""
        ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="vacaciones.xlsx",
        as_attachment=True
    )
