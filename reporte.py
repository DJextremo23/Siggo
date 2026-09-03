from flask import Blueprint, render_template, request, send_file, session, redirect
from io import BytesIO
from conexion import conexion
from utils import acceso_no_autorizado
from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from estilos_reporte import (
    COLOR_PRIMARIO, COLOR_ACENTO,
    estilos_pdf, build_pdf_tabla,
    configurar_encabezado_excel, aplicar_estilo_datos_excel
)

# Blueprint para reportes del admin - resumen, detalle y vacaciones con exportación a PDF y Excel
reporte_bp = Blueprint("reporte_bp", __name__)


# ==========================
# UTILIDADES
# ==========================
# Filtro genérico de texto sobre cualquier campo de una lista de diccionarios
def filtrar_por_texto(data, texto):
    """Filtra una lista de diccionarios buscando el texto en cualquier campo (case-insensitive)."""
    if not texto:
        return data
    texto_lower = texto.lower()
    return [row for row in data if texto_lower in ' '.join(str(v) for v in row.values()).lower()]


# ==========================
# UTILIDAD (FILTRO BASE)
# ==========================
# Construye la cláusula WHERE y los parámetros para filtrar guardias por fecha y usuarios
def construir_filtro(anio=None, mes=None, fecha_desde=None, fecha_hasta=None, ids_usuarios=None, alias_fecha="g.fecha_guardia"):
    filtro = "WHERE 1=1"
    params = []

    if fecha_desde:
        filtro += f" AND {alias_fecha} >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        filtro += f" AND {alias_fecha} <= %s"
        params.append(fecha_hasta)

    if not fecha_desde and not fecha_hasta:
        if anio:
            filtro += f" AND YEAR({alias_fecha}) = %s"
            params.append(int(anio))
        if mes:
            filtro += f" AND MONTH({alias_fecha}) = %s"
            params.append(int(mes))

    if ids_usuarios:
        placeholders = ",".join(["%s"] * len(ids_usuarios))
        filtro += f" AND u.id_usuario IN ({placeholders})"
        params.extend(ids_usuarios)

    return filtro, params


# ==========================
# REPORTE WEB
# ==========================
# Página principal de reportes: consulta resumen, detalle, vacaciones y resumen de vacaciones
@reporte_bp.route("/reportes")
def reporte():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar = request.args.get("buscar", "").strip()
    buscar_detalle = request.args.get("buscar_detalle", "").strip()
    buscar_vac = request.args.get("buscar_vac", "").strip()
    buscar_resumen_vac = request.args.get("buscar_resumen_vac", "").strip()

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT u.id_usuario, u.nombre, u.apellidos
            FROM usuarios u
            INNER JOIN usuarios_roles ur ON u.id_usuario = ur.id_usuario
            INNER JOIN roles r ON ur.id_rol = r.id_rol
            WHERE r.nombre_rol = 'fiscalizador' AND u.estado = 'activo'
            ORDER BY u.nombre, u.apellidos
        """)
        usuarios = cursor.fetchall()

        filtro, params = construir_filtro(
            anio=anio, mes=mes,
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            ids_usuarios=ids_usuarios
        )

        # ==========================
        # RESUMEN
        # ==========================
        cursor.execute(f"""
            SELECT 
                u.id_usuario,
                CONCAT(u.nombre,' ',u.apellidos) AS fiscalizador,
                COUNT(g.id_guardia) AS total_guardias,
                SUM(CASE WHEN f.id_feriado IS NOT NULL THEN 1 ELSE 0 END) AS guardias_feriado,
                COUNT(c.id_compensacion) AS compensaciones,
                (COUNT(g.id_guardia) - COUNT(c.id_compensacion)) AS pendientes
            FROM guardias g
            LEFT JOIN usuarios u ON g.id_usuario = u.id_usuario
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            {filtro}
            GROUP BY u.id_usuario
        """, params)

        reporte_data = cursor.fetchall()

        # ==========================
        # DETALLE
        # ==========================
        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre,' ',u.apellidos) AS fiscalizador,
                g.fecha_guardia,
                COALESCE(f.descripcion, '—') AS feriado_descripcion,
                CASE 
                    WHEN a.estado = 'asistio' THEN '✔ Asistió'
                    WHEN a.estado = 'falta' THEN '❌ Falta'
                    WHEN a.estado = 'justificado' THEN '🟡 Justificado'
                    ELSE '—'
                END AS asistencia,
                c.fecha_compensacion,
                c.observacion,
                CASE 
                    WHEN a.estado = 'falta' THEN '❌ No cumple'
                    WHEN a.estado = 'justificado' THEN '🟡 Justificado'
                    WHEN a.estado = 'asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                    WHEN a.estado = 'asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                    ELSE '—'
                END AS estado_compensacion
            FROM guardias g
            LEFT JOIN usuarios u ON g.id_usuario = u.id_usuario
            LEFT JOIN asistencia a ON g.id_guardia = a.id_guardia
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            {filtro}
            ORDER BY g.fecha_guardia DESC
        """, params)

        detalle = cursor.fetchall()

        # ==========================
        # VACACIONES
        # ==========================
        filtro_vac = "WHERE 1=1"
        params_vac = []

        if fecha_desde:
            filtro_vac += " AND v.fecha_inicio >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_fin <= %s"
            params_vac.append(fecha_hasta)

        if not fecha_desde and not fecha_hasta:
            if anio:
                filtro_vac += " AND YEAR(v.fecha_inicio) = %s"
                params_vac.append(int(anio))
            if mes:
                filtro_vac += " AND MONTH(v.fecha_inicio) = %s"
                params_vac.append(int(mes))

        if ids_usuarios:
            placeholders = ",".join(["%s"] * len(ids_usuarios))
            filtro_vac += f" AND v.id_usuario IN ({placeholders})"
            params_vac.extend(ids_usuarios)

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre, ' ', u.apellidos) AS fiscalizador,
                v.fecha_inicio,
                v.fecha_fin,
                DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1 AS dias_tomados,
                CASE
                    WHEN CURDATE() < v.fecha_inicio THEN 'pendiente'
                    WHEN CURDATE() BETWEEN v.fecha_inicio AND v.fecha_fin THEN 'en_curso'
                    ELSE 'finalizado'
                END AS estado
            FROM vacaciones v
            JOIN usuarios u ON u.id_usuario = v.id_usuario
            {filtro_vac}
            ORDER BY v.fecha_inicio DESC
        """, params_vac)

        vacaciones = cursor.fetchall()

        # ==========================
        # RESUMEN VACACIONES
        # ==========================
        cursor.execute(f"""
            SELECT
                CONCAT(u.nombre, ' ', u.apellidos) AS fiscalizador,
                SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1) AS dias_tomados,
                GREATEST(0, 30 - SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1)) AS dias_pendientes,
                GREATEST(0,
                    (TIMESTAMPDIFF(YEAR, u.fecha_ingreso, CURDATE()) * 30
                     - COALESCE((
                         SELECT SUM(DATEDIFF(v2.fecha_fin, v2.fecha_inicio) + 1)
                         FROM vacaciones v2
                         WHERE v2.id_usuario = u.id_usuario
                     ), 0))
                    - GREATEST(0, 30 - SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1))
                ) AS dias_pendientes_anteriores
            FROM vacaciones v
            JOIN usuarios u ON u.id_usuario = v.id_usuario
            {filtro_vac}
            GROUP BY u.id_usuario, u.nombre, u.apellidos
            ORDER BY u.nombre, u.apellidos
        """, params_vac)

        resumen_vacaciones = cursor.fetchall()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    reporte_data = filtrar_por_texto(reporte_data, buscar)
    detalle = filtrar_por_texto(detalle, buscar_detalle)
    vacaciones = filtrar_por_texto(vacaciones, buscar_vac)
    resumen_vacaciones = filtrar_por_texto(resumen_vacaciones, buscar_resumen_vac)

    return render_template(
        "reporte.html",
        reporte=reporte_data,
        detalle=detalle,
        vacaciones=vacaciones,
        resumen_vacaciones=resumen_vacaciones,
        usuarios=usuarios
    )


# ==========================
# ESTILOS PDF / EXCEL
# ==========================

def _estilos_pdf():
    return estilos_pdf()


def _build_pdf_tabla(encabezados, filas, estilos, ancho_disponible, columnas_centradas=None):
    return build_pdf_tabla(encabezados, filas, estilos, ancho_disponible, columnas_centradas)


def _configurar_encabezado_excel(ws, columnas, titulo=None):
    return configurar_encabezado_excel(ws, columnas, titulo)


def _aplicar_estilo_datos_excel(ws, columnas, data_start):
    aplicar_estilo_datos_excel(ws, columnas, data_start)




# ==========================
# EXPORTAR PDF (RESUMEN)
# ==========================
# Genera y descarga un PDF con el resumen de guardias
@reporte_bp.route("/exportar/pdf")
def exportar_pdf():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar = request.args.get("buscar")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro, params = construir_filtro(
            anio=anio, mes=mes,
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            ids_usuarios=ids_usuarios
        )

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre,' ',u.apellidos) AS fiscalizador,
                COUNT(g.id_guardia) AS total_guardias,
                SUM(CASE WHEN f.id_feriado IS NOT NULL THEN 1 ELSE 0 END) AS guardias_feriado,
                COUNT(c.id_compensacion) AS compensaciones,
                (COUNT(g.id_guardia) - COUNT(c.id_compensacion)) AS pendientes
            FROM guardias g
            LEFT JOIN usuarios u ON g.id_usuario = u.id_usuario
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            {filtro}
            GROUP BY u.id_usuario
        """, params)

        data = filtrar_por_texto(cursor.fetchall(), buscar)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elementos = []

    # Título con acento azul (simulado con una línea)
    elementos.append(Paragraph("Reporte de Guardias", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    # Subtítulo con filtros aplicados
    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Fiscalizador", "Total de Guardias", "Total de Feriados",
                "Total Compensaciones", "Compensaciones Pendientes"]
    filas = [[
        d["fiscalizador"],
        str(d["total_guardias"]),
        str(d["guardias_feriado"]),
        str(d["compensaciones"]),
        str(d["pendientes"])
    ] for d in data]
    ancho = landscape(A4)[0] - 60

    table = _build_pdf_tabla(columnas, filas, estilos, ancho, columnas_centradas=[1, 2, 3, 4])
    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, download_name="reporte.pdf", as_attachment=True)


# ==========================
# EXPORTAR EXCEL (RESUMEN)
# ==========================
# Genera y descarga un archivo Excel con el resumen de guardias
@reporte_bp.route("/exportar/excel")
def exportar_excel():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar = request.args.get("buscar")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro, params = construir_filtro(
            anio=anio, mes=mes,
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            ids_usuarios=ids_usuarios
        )

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre,' ',u.apellidos) AS fiscalizador,
                COUNT(g.id_guardia) AS total_guardias,
                SUM(CASE WHEN f.id_feriado IS NOT NULL THEN 1 ELSE 0 END) AS guardias_feriado,
                COUNT(c.id_compensacion) AS compensaciones,
                (COUNT(g.id_guardia) - COUNT(c.id_compensacion)) AS pendientes
            FROM guardias g
            LEFT JOIN usuarios u ON g.id_usuario = u.id_usuario
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            {filtro}
            GROUP BY u.id_usuario
        """, params)

        data = filtrar_por_texto(cursor.fetchall(), buscar)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    columnas = ["Fiscalizador", "Total de Guardias", "Total de Feriados",
                "Total Compensaciones", "Compensaciones Pendientes"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Reporte de Guardias")
    for d in data:
        ws.append([
            d["fiscalizador"],
            d["total_guardias"],
            d["guardias_feriado"],
            d["compensaciones"],
            d["pendientes"]
        ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)
    # Alinear centro columnas numéricas
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=2, max_col=5):
        for cel in row:
            cel.alignment = Alignment(horizontal='center', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="reporte.xlsx", as_attachment=True)


# ==========================
# EXPORTAR PDF DETALLE
# ==========================
# Genera y descarga un PDF con el detalle de guardias por fecha
@reporte_bp.route("/exportar/detalle_fecha/pdf")
def exportar_detalle_fecha_pdf():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar = request.args.get("buscar")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro, params = construir_filtro(
            anio=anio, mes=mes,
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            ids_usuarios=ids_usuarios
        )

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre,' ',u.apellidos) AS fiscalizador,
                g.fecha_guardia,
                COALESCE(f.descripcion, '—') AS feriado_descripcion,
                CASE 
                    WHEN a.estado = 'asistio' THEN '✔ Asistió'
                    WHEN a.estado = 'falta' THEN '❌ Falta'
                    WHEN a.estado = 'justificado' THEN '🟡 Justificado'
                    ELSE '—'
                END AS asistencia,
                c.fecha_compensacion,
                c.observacion,
                CASE 
                    WHEN a.estado = 'falta' THEN '❌ No cumple'
                    WHEN a.estado = 'justificado' THEN '🟡 Justificado'
                    WHEN a.estado = 'asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                    WHEN a.estado = 'asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                    ELSE '—'
                END AS estado_compensacion
            FROM guardias g
            LEFT JOIN usuarios u ON g.id_usuario = u.id_usuario
            LEFT JOIN asistencia a ON g.id_guardia = a.id_guardia
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            {filtro}
            ORDER BY g.fecha_guardia DESC
        """, params)

        detalle = filtrar_por_texto(cursor.fetchall(), buscar)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=25, rightMargin=25, topMargin=30, bottomMargin=30)
    elementos = []

    elementos.append(Paragraph("Detalle de Guardias", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Fiscalizador", "Fecha", "Feriado", "Asistencia", "Compensación", "Estado", "Observaciones"]
    filas = [[
        d["fiscalizador"],
        str(d["fecha_guardia"]) if d["fecha_guardia"] else "",
        d["feriado_descripcion"],
        d["asistencia"],
        str(d["fecha_compensacion"]) if d["fecha_compensacion"] else "\u2014",
        d["estado_compensacion"],
        d["observacion"] or "\u2014"
    ] for d in detalle]
    ancho = landscape(A4)[0] - 50

    table = _build_pdf_tabla(columnas, filas, estilos, ancho)
    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, download_name="detalle_fecha.pdf", as_attachment=True)


# ==========================
# EXPORTAR EXCEL DETALLE
# ==========================
# Genera y descarga un archivo Excel con el detalle de guardias por fecha
@reporte_bp.route("/exportar/detalle_fecha/excel")
def exportar_detalle_fecha_excel():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar = request.args.get("buscar")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro, params = construir_filtro(
            anio=anio, mes=mes,
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            ids_usuarios=ids_usuarios
        )

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre,' ',u.apellidos) AS fiscalizador,
                g.fecha_guardia,
                COALESCE(f.descripcion, '—') AS feriado_descripcion,
                CASE 
                    WHEN a.estado = 'asistio' THEN '✔ Asistió'
                    WHEN a.estado = 'falta' THEN '❌ Falta'
                    WHEN a.estado = 'justificado' THEN '🟡 Justificado'
                    ELSE '—'
                END AS asistencia,
                c.fecha_compensacion,
                c.observacion,
                CASE 
                    WHEN a.estado = 'falta' THEN '❌ No cumple'
                    WHEN a.estado = 'justificado' THEN '🟡 Justificado'
                    WHEN a.estado = 'asistio' AND c.id_compensacion IS NULL THEN '⚠️ Pendiente'
                    WHEN a.estado = 'asistio' AND c.id_compensacion IS NOT NULL THEN '✔ Compensado'
                    ELSE '—'
                END AS estado_compensacion
            FROM guardias g
            LEFT JOIN usuarios u ON g.id_usuario = u.id_usuario
            LEFT JOIN asistencia a ON g.id_guardia = a.id_guardia
            LEFT JOIN compensaciones c ON g.id_guardia = c.id_guardia
            LEFT JOIN feriados f ON g.id_feriado = f.id_feriado
            {filtro}
            ORDER BY g.fecha_guardia DESC
        """, params)

        detalle = filtrar_por_texto(cursor.fetchall(), buscar)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Fecha"
    columnas = ["Fiscalizador", "Fecha", "Feriado", "Asistencia", "Compensación", "Estado", "Observaciones"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Detalle de Guardias")
    for d in detalle:
        ws.append([
            d["fiscalizador"],
            str(d["fecha_guardia"]) if d["fecha_guardia"] else "",
            d["feriado_descripcion"],
            d["asistencia"],
            str(d["fecha_compensacion"]) if d["fecha_compensacion"] else "",
            d["estado_compensacion"],
            d["observacion"] or ""
        ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="detalle_fecha.xlsx", as_attachment=True)


# ==========================
# EXPORTAR PDF VACACIONES
# ==========================
# Genera y descarga un PDF con el detalle de vacaciones
@reporte_bp.route("/exportar/vacaciones/pdf")
def exportar_vacaciones_pdf():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar_vac = request.args.get("buscar_vac")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_vac = "WHERE 1=1"
        params_vac = []

        if fecha_desde:
            filtro_vac += " AND v.fecha_inicio >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_fin <= %s"
            params_vac.append(fecha_hasta)

        if not fecha_desde and not fecha_hasta:
            if anio:
                filtro_vac += " AND YEAR(v.fecha_inicio) = %s"
                params_vac.append(int(anio))
            if mes:
                filtro_vac += " AND MONTH(v.fecha_inicio) = %s"
                params_vac.append(int(mes))

        if ids_usuarios:
            placeholders = ",".join(["%s"] * len(ids_usuarios))
            filtro_vac += f" AND v.id_usuario IN ({placeholders})"
            params_vac.extend(ids_usuarios)

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre, ' ', u.apellidos) AS fiscalizador,
                v.fecha_inicio,
                v.fecha_fin,
                DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1 AS dias_tomados,
                CASE
                    WHEN CURDATE() < v.fecha_inicio THEN 'pendiente'
                    WHEN CURDATE() BETWEEN v.fecha_inicio AND v.fecha_fin THEN 'en_curso'
                    ELSE 'finalizado'
                END AS estado
            FROM vacaciones v
            JOIN usuarios u ON u.id_usuario = v.id_usuario
            {filtro_vac}
            ORDER BY v.fecha_inicio DESC
        """, params_vac)

        data = filtrar_por_texto(cursor.fetchall(), buscar_vac)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elementos = []

    elementos.append(Paragraph("Detalle de Vacaciones", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Fiscalizador", "Inicio", "Fin", "Días Tomados", "Estado"]
    filas = [[
        v["fiscalizador"],
        str(v["fecha_inicio"]) if v["fecha_inicio"] else "",
        str(v["fecha_fin"]) if v["fecha_fin"] else "",
        str(v["dias_tomados"]),
        v["estado"].replace('_', ' ').title()
    ] for v in data]
    ancho = landscape(A4)[0] - 60

    table = _build_pdf_tabla(columnas, filas, estilos, ancho, columnas_centradas=[3])
    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, download_name="vacaciones.pdf", as_attachment=True)


# ==========================
# EXPORTAR EXCEL VACACIONES
# ==========================
# Genera y descarga un archivo Excel con el detalle de vacaciones
@reporte_bp.route("/exportar/vacaciones/excel")
def exportar_vacaciones_excel():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar_vac = request.args.get("buscar_vac")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_vac = "WHERE 1=1"
        params_vac = []

        if fecha_desde:
            filtro_vac += " AND v.fecha_inicio >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_fin <= %s"
            params_vac.append(fecha_hasta)

        if not fecha_desde and not fecha_hasta:
            if anio:
                filtro_vac += " AND YEAR(v.fecha_inicio) = %s"
                params_vac.append(int(anio))
            if mes:
                filtro_vac += " AND MONTH(v.fecha_inicio) = %s"
                params_vac.append(int(mes))

        if ids_usuarios:
            placeholders = ",".join(["%s"] * len(ids_usuarios))
            filtro_vac += f" AND v.id_usuario IN ({placeholders})"
            params_vac.extend(ids_usuarios)

        cursor.execute(f"""
            SELECT 
                CONCAT(u.nombre, ' ', u.apellidos) AS fiscalizador,
                v.fecha_inicio,
                v.fecha_fin,
                DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1 AS dias_tomados,
                CASE
                    WHEN CURDATE() < v.fecha_inicio THEN 'pendiente'
                    WHEN CURDATE() BETWEEN v.fecha_inicio AND v.fecha_fin THEN 'en_curso'
                    ELSE 'finalizado'
                END AS estado
            FROM vacaciones v
            JOIN usuarios u ON u.id_usuario = v.id_usuario
            {filtro_vac}
            ORDER BY v.fecha_inicio DESC
        """, params_vac)

        data = filtrar_por_texto(cursor.fetchall(), buscar_vac)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vacaciones"
    columnas = ["Fiscalizador", "Inicio", "Fin", "Días Tomados", "Estado"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Detalle de Vacaciones")
    for v in data:
        ws.append([
            v["fiscalizador"],
            str(v["fecha_inicio"]) if v["fecha_inicio"] else "",
            str(v["fecha_fin"]) if v["fecha_fin"] else "",
            v["dias_tomados"],
            v["estado"].replace('_', ' ').title()
        ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="vacaciones.xlsx", as_attachment=True)


# ==========================
# EXPORTAR PDF RESUMEN VACACIONES
# ==========================
# Genera y descarga un PDF con el resumen de vacaciones por fiscalizador
@reporte_bp.route("/exportar/resumen_vacaciones/pdf")
def exportar_resumen_vacaciones_pdf():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar_resumen_vac = request.args.get("buscar_resumen_vac")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_vac = "WHERE 1=1"
        params_vac = []

        if fecha_desde:
            filtro_vac += " AND v.fecha_inicio >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_fin <= %s"
            params_vac.append(fecha_hasta)

        if not fecha_desde and not fecha_hasta:
            if anio:
                filtro_vac += " AND YEAR(v.fecha_inicio) = %s"
                params_vac.append(int(anio))
            if mes:
                filtro_vac += " AND MONTH(v.fecha_inicio) = %s"
                params_vac.append(int(mes))

        if ids_usuarios:
            placeholders = ",".join(["%s"] * len(ids_usuarios))
            filtro_vac += f" AND v.id_usuario IN ({placeholders})"
            params_vac.extend(ids_usuarios)

        cursor.execute(f"""
            SELECT
                CONCAT(u.nombre, ' ', u.apellidos) AS fiscalizador,
                SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1) AS dias_tomados,
                GREATEST(0, 30 - SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1)) AS dias_pendientes,
                GREATEST(0,
                    (TIMESTAMPDIFF(YEAR, u.fecha_ingreso, CURDATE()) * 30
                     - COALESCE((
                         SELECT SUM(DATEDIFF(v2.fecha_fin, v2.fecha_inicio) + 1)
                         FROM vacaciones v2
                         WHERE v2.id_usuario = u.id_usuario
                     ), 0))
                    - GREATEST(0, 30 - SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1))
                ) AS dias_pendientes_anteriores
            FROM vacaciones v
            JOIN usuarios u ON u.id_usuario = v.id_usuario
            {filtro_vac}
            GROUP BY u.id_usuario, u.nombre, u.apellidos
            ORDER BY u.nombre, u.apellidos
        """, params_vac)

        data = filtrar_por_texto(cursor.fetchall(), buscar_resumen_vac)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    estilos = _estilos_pdf()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elementos = []

    elementos.append(Paragraph("Resumen de Vacaciones", estilos['titulo']))
    elementos.append(Spacer(1, 2))
    elementos.append(HRFlowable(width='40', thickness=3, color=colors.HexColor(COLOR_ACENTO),
                                 spaceAfter=10, hAlign='LEFT'))

    partes = []
    if anio: partes.append(f"Año: {anio}")
    if fecha_desde: partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta: partes.append(f"Hasta: {fecha_hasta}")
    if partes:
        elementos.append(Paragraph(" | ".join(partes), estilos['subtitulo']))

    columnas = ["Fiscalizador", "Días Tomados", "Días Pendientes", "Días Pend. Años Anteriores"]
    filas = [[
        r["fiscalizador"],
        str(r["dias_tomados"]),
        str(r["dias_pendientes"]),
        str(r["dias_pendientes_anteriores"])
    ] for r in data]
    ancho = landscape(A4)[0] - 60

    table = _build_pdf_tabla(columnas, filas, estilos, ancho, columnas_centradas=[1, 2, 3])
    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, download_name="resumen_vacaciones.pdf", as_attachment=True)


# ==========================
# EXPORTAR EXCEL RESUMEN VACACIONES
# ==========================
# Genera y descarga un archivo Excel con el resumen de vacaciones por fiscalizador
@reporte_bp.route("/exportar/resumen_vacaciones/excel")
def exportar_resumen_vacaciones_excel():
    if "usuario" not in session:
        return redirect("/")
    if session.get("perfil_activo") != "admin":
        return acceso_no_autorizado()

    anio = request.args.get("anio")
    mes = request.args.get("mes")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    ids_usuarios = request.args.getlist("id_usuario")
    buscar_resumen_vac = request.args.get("buscar_resumen_vac")

    conn = None
    cursor = None
    try:
        conn = conexion()
        cursor = conn.cursor(dictionary=True)

        filtro_vac = "WHERE 1=1"
        params_vac = []

        if fecha_desde:
            filtro_vac += " AND v.fecha_inicio >= %s"
            params_vac.append(fecha_desde)

        if fecha_hasta:
            filtro_vac += " AND v.fecha_fin <= %s"
            params_vac.append(fecha_hasta)

        if not fecha_desde and not fecha_hasta:
            if anio:
                filtro_vac += " AND YEAR(v.fecha_inicio) = %s"
                params_vac.append(int(anio))
            if mes:
                filtro_vac += " AND MONTH(v.fecha_inicio) = %s"
                params_vac.append(int(mes))

        if ids_usuarios:
            placeholders = ",".join(["%s"] * len(ids_usuarios))
            filtro_vac += f" AND v.id_usuario IN ({placeholders})"
            params_vac.extend(ids_usuarios)

        cursor.execute(f"""
            SELECT
                CONCAT(u.nombre, ' ', u.apellidos) AS fiscalizador,
                SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1) AS dias_tomados,
                GREATEST(0, 30 - SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1)) AS dias_pendientes,
                GREATEST(0,
                    (TIMESTAMPDIFF(YEAR, u.fecha_ingreso, CURDATE()) * 30
                     - COALESCE((
                         SELECT SUM(DATEDIFF(v2.fecha_fin, v2.fecha_inicio) + 1)
                         FROM vacaciones v2
                         WHERE v2.id_usuario = u.id_usuario
                     ), 0))
                    - GREATEST(0, 30 - SUM(DATEDIFF(v.fecha_fin, v.fecha_inicio) + 1))
                ) AS dias_pendientes_anteriores
            FROM vacaciones v
            JOIN usuarios u ON u.id_usuario = v.id_usuario
            {filtro_vac}
            GROUP BY u.id_usuario, u.nombre, u.apellidos
            ORDER BY u.nombre, u.apellidos
        """, params_vac)

        data = filtrar_por_texto(cursor.fetchall(), buscar_resumen_vac)
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Vacaciones"
    columnas = ["Fiscalizador", "Días Tomados", "Días Pendientes", "Días Pend. Años Anteriores"]
    data_start = _configurar_encabezado_excel(ws, columnas, titulo="Resumen de Vacaciones")
    for r in data:
        ws.append([
            r["fiscalizador"],
            r["dias_tomados"],
            r["dias_pendientes"],
            r["dias_pendientes_anteriores"]
        ])
    _aplicar_estilo_datos_excel(ws, columnas, data_start)
    # Center numeric columns
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=2, max_col=4):
        for cel in row:
            cel.alignment = Alignment(horizontal='center', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="resumen_vacaciones.xlsx", as_attachment=True)
